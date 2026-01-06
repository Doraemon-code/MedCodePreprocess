import streamlit as st
import pandas as pd
import html
import io
import time
import logging
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from app.ai_extractor import ai_extract_batch
from app.config_store import load_all_configs, save_current_config, load_config, delete_config
from app.rules import evaluate_condition, extract_value, process_variable_rules
from app.settings import AI_CONFIG

# ==================== 日志配置 ====================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('log.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

MAX_UI_LOG_LINES = 200


class UILogHandler(logging.Handler):
    def emit(self, record):
        try:
            message = self.format(record)
            logs = st.session_state.get("ui_logs")
            if logs is None:
                st.session_state.ui_logs = []
                logs = st.session_state.ui_logs
            logs.append(message)
            if len(logs) > MAX_UI_LOG_LINES:
                del logs[: len(logs) - MAX_UI_LOG_LINES]
        except Exception:
            pass


root_logger = logging.getLogger()
if not any(getattr(h, "name", "") == "ui_log_handler" for h in root_logger.handlers):
    ui_handler = UILogHandler()
    ui_handler.name = "ui_log_handler"
    ui_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    root_logger.addHandler(ui_handler)

logger.info("=" * 80)
logger.info("程序启动")
logger.info("=" * 80)

# ==================== AI配置 ====================
logger.info(f"AI配置加载完成: MODEL={AI_CONFIG['MODEL']}, BASE_URL={AI_CONFIG['BASE_URL']}")

# 页面配置
st.set_page_config(
    page_title="医学编码数据预处理器",
    page_icon="📊",
    layout="wide"
)

logger.info("Streamlit页面配置完成")

# 自定义CSS样式
st.markdown("""
<style>
    .main {
        background: linear-gradient(135deg, #e0f2fe 0%, #ddd6fe 100%);
    }
    .stButton>button {
        border-radius: 0.5rem;
        font-weight: 600;
        transition: all 0.3s;
    }
    h1 {
        color: #1f2937;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .subtitle {
        text-align: center;
        color: #6b7280;
        margin-bottom: 2rem;
    }
    .section-header {
        background: linear-gradient(90deg, #4f46e5 0%, #7c3aed 100%);
        color: white;
        padding: 0.75rem 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0 0.5rem 0;
        font-weight: 600;
    }
    .rule-card {
        background: white;
        border-left: 4px solid #4f46e5;
        padding: 1rem;
        margin: 0.5rem 0;
        border-radius: 0.5rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    }
    .variable-header {
        background: #f3f4f6;
        padding: 0.75rem;
        border-radius: 0.5rem;
        margin: 1rem 0 0.5rem 0;
        font-weight: 600;
        color: #1f2937;
    }
    .rule-summary {
        background: #f9fafb;
        padding: 0.5rem;
        margin: 0.25rem 0;
        border-radius: 0.25rem;
        font-family: monospace;
        font-size: 0.9rem;
        color: #374151;
    }
    .log-panel {
        max-height: 340px;
        min-height: 260px;
        overflow-y: auto;
        background: rgba(255, 255, 255, 0.85);
        color: #1f2937;
        font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;
        font-size: 0.8rem;
        padding: 0.75rem;
        border-radius: 0.5rem;
        border: 1px solid rgba(79, 70, 229, 0.18);
        box-shadow: 0 6px 16px rgba(31, 41, 55, 0.08);
        white-space: pre-wrap;
    }
    @media (max-width: 768px) {
        .log-panel {
            max-height: 260px;
            min-height: 200px;
        }
    }
</style>
""", unsafe_allow_html=True)

# 初始化session state
if 'uploaded_file' not in st.session_state:
    st.session_state.uploaded_file = None
    logger.info("初始化 session_state: uploaded_file")
if 'excel_data' not in st.session_state:
    st.session_state.excel_data = None
    logger.info("初始化 session_state: excel_data")
if 'selected_sheets' not in st.session_state:
    st.session_state.selected_sheets = {}
    logger.info("初始化 session_state: selected_sheets")
if 'sheet_variables' not in st.session_state:
    st.session_state.sheet_variables = {}
    logger.info("初始化 session_state: sheet_variables")
if 'ai_cache' not in st.session_state:
    st.session_state.ai_cache = {}
    logger.info("初始化 session_state: ai_cache")
if 'ui_logs' not in st.session_state:
    st.session_state.ui_logs = []
    logger.info("初始化 session_state: ui_logs")


def render_log_panel(placeholder):
    logs = st.session_state.get("ui_logs", [])
    if logs:
        content = "\n".join(logs[-MAX_UI_LOG_LINES:])
    else:
        content = "暂无日志"
    placeholder.markdown(
        f"<div class='log-panel'>{html.escape(content)}</div>",
        unsafe_allow_html=True,
    )


# ==================== 侧边栏：配置管理 ====================

with st.sidebar:
    st.markdown("### 🤖 AI设置")
    model_options = ["deepseek-chat", "deepseek-reasoner"]
    default_model = AI_CONFIG["MODEL"] if AI_CONFIG["MODEL"] in model_options else model_options[0]
    if "ai_model" not in st.session_state:
        st.session_state.ai_model = default_model
    selected_model = st.selectbox(
        "模型",
        options=model_options,
        index=model_options.index(st.session_state.ai_model),
        key="ai_model",
    )
    if selected_model != AI_CONFIG["MODEL"]:
        AI_CONFIG["MODEL"] = selected_model
        logger.info("AI模型切换为: %s", selected_model)

    st.markdown("### 💾 配置管理")
    
    with st.expander("保存当前配置", expanded=False):
        save_name = st.text_input(
            "配置名称",
            placeholder="例如: 默认配置",
            key="save_config_name"
        )
        if st.button("💾 保存", key="save_config_btn", use_container_width=True):
            if save_name:
                if save_current_config(save_name):
                    st.success(f"✅ 配置 '{save_name}' 已保存!")
                    st.rerun()
            else:
                st.warning("⚠️ 请输入配置名称")
    
    with st.expander("加载配置", expanded=False):
        all_configs = load_all_configs()
        if all_configs:
            config_options = list(all_configs.keys())
            selected_config = st.selectbox(
                "选择配置",
                options=config_options,
                key="load_config_select"
            )
            
            if selected_config:
                saved_time = all_configs[selected_config].get('saved_time', '未知')
                st.caption(f"⏰ {saved_time}")
                
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("📥 加载", key="load_btn", use_container_width=True):
                        if load_config(selected_config):
                            st.success(f"✅ 已加载 '{selected_config}'")
                            st.rerun()
                
                with col2:
                    if st.button("🗑️ 删除", key="delete_btn", use_container_width=True):
                        if delete_config(selected_config):
                            st.success(f"✅ 已删除 '{selected_config}'")
                            st.rerun()
        else:
            st.info("ℹ️ 暂无保存的配置")

# ==================== 主页面 ====================

st.markdown("<h1>📊 医学编码数据预处理器</h1>", unsafe_allow_html=True)
st.markdown("<p class='subtitle'>导入、配置、导出 - 轻松处理您的数据（含AI提取）</p>", unsafe_allow_html=True)

# ==================== 布局：上传区域 + 日志面板 ====================

col_upload, col_log = st.columns([1, 1])

with col_upload:
    st.markdown("<div class='section-header'>📁 上传 Excel 文件</div>", unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "选择Excel文件",
        type=['xlsx', 'xls'],
        help="支持 .xlsx 和 .xls 格式",
        label_visibility="collapsed"
    )

    if uploaded_file is not None:
        logger.info(f"用户上传文件: {uploaded_file.name}")
        try:
            excel_file = pd.ExcelFile(uploaded_file)
            st.session_state.uploaded_file = uploaded_file
            st.session_state.excel_data = excel_file

            logger.info(f"Excel文件读取成功: {len(excel_file.sheet_names)} 个工作表")
            logger.info(f"工作表列表: {excel_file.sheet_names}")

            if not st.session_state.selected_sheets:
                st.session_state.selected_sheets = {
                    sheet: True for sheet in excel_file.sheet_names
                }
                logger.info("默认选中所有工作表")

            st.success(f"✅ 成功加载: {uploaded_file.name} ({len(excel_file.sheet_names)} 个工作表)")

        except Exception as e:
            logger.error(f"文件读取失败: {str(e)}", exc_info=True)
            st.error(f"❌ 文件读取失败: {str(e)}")

with col_log:
    st.markdown("<div class='section-header'>🧾 实时日志</div>", unsafe_allow_html=True)
    log_panel_placeholder = st.empty()
    render_log_panel(log_panel_placeholder)

st.markdown("---")

# ==================== 主要区域：Sheet选择 + 配置 ====================

if st.session_state.excel_data is not None:
    
    col_sheets, col_config = st.columns([1, 3])
    
    with col_sheets:
        st.markdown("<div class='section-header'>📋 选择工作表</div>", unsafe_allow_html=True)
        
        if 'sheet_select_trigger' not in st.session_state:
            st.session_state.sheet_select_trigger = 0
        
        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("✅ 全选", use_container_width=True):
                logger.info("用户点击全选")
                for sheet in st.session_state.selected_sheets:
                    st.session_state.selected_sheets[sheet] = True
                st.session_state.sheet_select_trigger += 1
                st.rerun()
        with col_b:
            if st.button("❌ 全不选", use_container_width=True):
                logger.info("用户点击全不选")
                for sheet in st.session_state.selected_sheets:
                    st.session_state.selected_sheets[sheet] = False
                st.session_state.sheet_select_trigger += 1
                st.rerun()
        
        st.markdown("---")
        
        for sheet_name in st.session_state.excel_data.sheet_names:
            checked = st.checkbox(
                f"📄 {sheet_name}",
                value=st.session_state.selected_sheets.get(sheet_name, True),
                key=f"sheet_{sheet_name}_{st.session_state.sheet_select_trigger}"
            )
            st.session_state.selected_sheets[sheet_name] = checked
    
    with col_config:
        st.markdown("<div class='section-header'>⚙️ 变量配置</div>", unsafe_allow_html=True)
        
        selected_sheets = [name for name, sel in st.session_state.selected_sheets.items() if sel]
        logger.info(f"当前选中工作表数: {len(selected_sheets)}")
        
        if not selected_sheets:
            st.warning("⚠️ 请先选择至少一个工作表")
        else:
            for sheet_name in selected_sheets:
                with st.expander(f"📊 {sheet_name}", expanded=True):
                    
                    if sheet_name not in st.session_state.sheet_variables:
                        st.session_state.sheet_variables[sheet_name] = {}
                        logger.info(f"初始化工作表配置: {sheet_name}")
                    
                    sheet_vars = st.session_state.sheet_variables[sheet_name]
                    
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        new_var_name = st.text_input(
                            "新变量名",
                            placeholder="例如: ROUTE, INDICATION",
                            key=f"new_var_{sheet_name}"
                        )
                    with col2:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("➕ 添加变量", key=f"add_var_{sheet_name}"):
                            if new_var_name and new_var_name not in sheet_vars:
                                sheet_vars[new_var_name] = {
                                    'separator': ';',
                                    'rules': []
                                }
                                logger.info(f"添加新变量: {sheet_name}.{new_var_name}")
                                st.rerun()
                            elif new_var_name in sheet_vars:
                                st.warning("⚠️ 变量名已存在")
                            else:
                                st.warning("⚠️ 请输入变量名")
                    
                    st.markdown("---")
                    
                    for var_name in list(sheet_vars.keys()):
                        var_config = sheet_vars[var_name]
                        
                        st.markdown(f"<div class='variable-header'>📋 {var_name}</div>", unsafe_allow_html=True)
                        
                        col1, col2, col3 = st.columns([2, 2, 1])
                        with col1:
                            var_config['separator'] = st.text_input(
                                "分隔符",
                                value=var_config.get('separator', ';'),
                                key=f"sep_{sheet_name}_{var_name}"
                            )
                        with col2:
                            st.markdown("<br>", unsafe_allow_html=True)
                            if st.button(f"➕ 添加规则", key=f"add_rule_{sheet_name}_{var_name}"):
                                var_config['rules'].append({
                                    'condition_column': '',
                                    'condition_operator': '=',
                                    'condition_value': '',
                                    'extract_type': '直接提取',
                                    'extract_value_type': '从列提取',
                                    'extract_value': '',
                                    'regex_pattern': '',
                                    'capture_group': 1
                                })
                                logger.info(f"添加规则: {sheet_name}.{var_name}")
                                st.rerun()
                        with col3:
                            st.markdown("<br>", unsafe_allow_html=True)
                            if st.button("🗑️", key=f"del_var_{sheet_name}_{var_name}"):
                                logger.info(f"删除变量: {sheet_name}.{var_name}")
                                del sheet_vars[var_name]
                                st.rerun()
                        
                        if var_config['rules']:
                            for idx, rule in enumerate(var_config['rules']):
                                cond_col = rule.get('condition_column', '')
                                cond_op = rule.get('condition_operator', '=')
                                cond_val = rule.get('condition_value', '')
                                ext_type = rule.get('extract_type', '直接提取')
                                ext_val_type = rule.get('extract_value_type', '从列提取')
                                ext_val = rule.get('extract_value', '')
                                
                                rule_text = f"{'├─' if idx < len(var_config['rules'])-1 else '└─'} 规则{idx+1}: "
                                rule_text += f"当 {cond_col} {cond_op} "
                                rule_text += f'"{cond_val}"' if cond_val else '(空)'
                                rule_text += f" 时，{ext_type} "
                                
                                if ext_type == "AI提取":
                                    rule_text += f"🤖 从 {ext_val} 提取成分"
                                elif ext_val_type == "固定文本":
                                    rule_text += f'"{ext_val}"'
                                else:
                                    rule_text += f'{ext_val}'
                                
                                if ext_type == "正则提取":
                                    regex = rule.get('regex_pattern', '')
                                    cap_grp = rule.get('capture_group', 1)
                                    rule_text += f" (模式: {regex}, 组{cap_grp})"
                                
                                st.markdown(f"<div class='rule-summary'>{rule_text}</div>", unsafe_allow_html=True)
                        
                        # 编辑规则
                        for idx, rule in enumerate(var_config['rules']):
                            with st.expander(f"🔧 规则 {idx + 1}", expanded=False):
                                
                                if st.button("🗑️ 删除此规则", key=f"del_rule_{sheet_name}_{var_name}_{idx}"):
                                    var_config['rules'].pop(idx)
                                    logger.info(f"删除规则: {sheet_name}.{var_name}.规则{idx+1}")
                                    st.rerun()
                                
                                st.markdown("**条件设置**")
                                col1, col2, col3 = st.columns(3)
                                
                                with col1:
                                    rule['condition_column'] = st.text_input(
                                        "判断变量(列名)",
                                        value=rule.get('condition_column', ''),
                                        placeholder="例如: CMROUTE",
                                        key=f"cond_col_{sheet_name}_{var_name}_{idx}"
                                    )
                                
                                with col2:
                                    operators = ["=", "<>", "包含", "不包含", ">", "<", ">=", "<="]
                                    current_op = rule.get('condition_operator', '=')
                                    rule['condition_operator'] = st.selectbox(
                                        "逻辑比较符",
                                        options=operators,
                                        index=operators.index(current_op) if current_op in operators else 0,
                                        key=f"cond_op_{sheet_name}_{var_name}_{idx}"
                                    )
                                
                                with col3:
                                    rule['condition_value'] = st.text_input(
                                        "判断值",
                                        value=rule.get('condition_value', ''),
                                        placeholder="留空表示空值",
                                        key=f"cond_val_{sheet_name}_{var_name}_{idx}"
                                    )
                                
                                st.markdown("---")
                                
                                st.markdown("**提取设置**")
                                
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    extract_types = ["直接提取", "正则提取", "AI提取"]
                                    current_ext = rule.get('extract_type', '直接提取')
                                    rule['extract_type'] = st.selectbox(
                                        "提取方式",
                                        options=extract_types,
                                        index=extract_types.index(current_ext) if current_ext in extract_types else 0,
                                        key=f"ext_type_{sheet_name}_{var_name}_{idx}",
                                        help="AI提取：使用DeepSeek AI从药物名称中提取核心成分"
                                    )
                                
                                with col2:
                                    value_types = ["从列提取", "固定文本"]
                                    current_val_type = rule.get('extract_value_type', '从列提取')
                                    rule['extract_value_type'] = st.selectbox(
                                        "提取值类型",
                                        options=value_types,
                                        index=value_types.index(current_val_type) if current_val_type in value_types else 0,
                                        key=f"ext_val_type_{sheet_name}_{var_name}_{idx}"
                                    )
                                
                                if rule['extract_type'] == "AI提取":
                                    rule['extract_value_type'] = "从列提取"
                                    rule['extract_value'] = st.text_input(
                                        "源数据列名 (AI将从此列提取药物成分)",
                                        value=rule.get('extract_value', ''),
                                        placeholder="例如: CMDECOD",
                                        key=f"ext_val_{sheet_name}_{var_name}_{idx}",
                                        help="AI会分析该列的药物名称并提取核心成分"
                                    )
                                    st.info("💡 AI提取会自动识别药物成分，无需正则表达式")
                                
                                elif rule['extract_value_type'] == "从列提取":
                                    rule['extract_value'] = st.text_input(
                                        "提取值(列名)",
                                        value=rule.get('extract_value', ''),
                                        placeholder="例如: CMROUTE",
                                        key=f"ext_val_{sheet_name}_{var_name}_{idx}"
                                    )
                                else:
                                    rule['extract_value'] = st.text_input(
                                        "提取值(固定文本)",
                                        value=rule.get('extract_value', ''),
                                        placeholder="例如: 预防感冒",
                                        key=f"ext_val_{sheet_name}_{var_name}_{idx}"
                                    )
                                
                                if rule['extract_type'] == "正则提取":
                                    col1, col2 = st.columns([3, 1])
                                    with col1:
                                        rule['regex_pattern'] = st.text_input(
                                            "正则表达式",
                                            value=rule.get('regex_pattern', ''),
                                            placeholder=r"例如: (\d+)#(.+?)[;,]",
                                            key=f"regex_{sheet_name}_{var_name}_{idx}",
                                            help="使用 .+? 进行非贪婪匹配"
                                        )
                                    with col2:
                                        rule['capture_group'] = st.number_input(
                                            "捕获组序号",
                                            value=rule.get('capture_group', 1),
                                            min_value=1,
                                            step=1,
                                            key=f"cap_grp_{sheet_name}_{var_name}_{idx}"
                                        )
                        
                        st.markdown("<br>", unsafe_allow_html=True)
    
    # ==================== 导出区域 ====================
    st.markdown("---")
    st.markdown("<div class='section-header'>📥 导出处理后的文件</div>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col2:
        if st.button("🚀 处理并导出", type="primary", use_container_width=True):
            logger.info("=" * 80)
            logger.info("开始处理并导出")
            logger.info("=" * 80)
            render_log_panel(log_panel_placeholder)
            
            try:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    
                    for sheet_name in selected_sheets:
                        logger.info(f"处理工作表: {sheet_name}")
                        
                        logger.info(f"  读取数据: {sheet_name}")
                        df = pd.read_excel(
                            st.session_state.uploaded_file,
                            sheet_name=sheet_name,
                            dtype=str
                        )
                        logger.info(f"  数据读取完成: 行数={len(df)}, 列数={len(df.columns)}")
                        
                        if sheet_name in st.session_state.sheet_variables:
                            sheet_vars = st.session_state.sheet_variables[sheet_name]
                            logger.info(f"  该工作表有 {len(sheet_vars)} 个变量需要处理")
                            
                            for var_name, var_config in sheet_vars.items():
                                logger.info(f"  处理变量: {var_name}")
                                
                                separator = var_config.get('separator', ';')
                                rules = var_config.get('rules', [])
                                logger.info(f"    规则数: {len(rules)}, 分隔符: '{separator}'")
                                
                                has_ai_rules = any(r.get('extract_type') == 'AI提取' for r in rules)
                                
                                if has_ai_rules:
                                    logger.info(f"    检测到AI提取规则")
                                    
                                    ai_tasks = []
                                    
                                    for row_idx, row in df.iterrows():
                                        for rule_idx, rule in enumerate(rules):
                                            if rule.get('extract_type') == 'AI提取':
                                                cond_col = rule.get('condition_column', '')
                                                cond_op = rule.get('condition_operator', '=')
                                                cond_val = rule.get('condition_value', '')
                                                
                                                if cond_col and cond_col in df.columns:
                                                    if evaluate_condition(row[cond_col], cond_op, cond_val):
                                                        source_col = rule.get('extract_value', '')
                                                        if source_col and source_col in df.columns:
                                                            ai_tasks.append((row_idx, rule_idx, source_col, row[source_col]))
                                    
                                    logger.info(f"    需要AI处理的任务数: {len(ai_tasks)}")
                                    
                                    from collections import defaultdict
                                    col_groups = defaultdict(list)
                                    for task in ai_tasks:
                                        row_idx, rule_idx, source_col, value = task
                                        col_groups[source_col].append((row_idx, value))
                                    
                                    ai_results = {}
                                    
                                    for source_col, tasks in col_groups.items():
                                        logger.info(f"    AI批量处理列 '{source_col}': {len(tasks)} 条数据")
                                        
                                        values = [v for _, v in tasks]
                                        row_indices = [idx for idx, _ in tasks]
                                        
                                        batch_size = AI_CONFIG["BATCH_SIZE"]
                                        total_batches = (len(values) + batch_size - 1) // batch_size
                                        
                                        with st.spinner(f"正在使用AI提取 {var_name} (列: {source_col}, 共{len(values)}条)..."):
                                            for batch_idx in range(total_batches):
                                                start_idx = batch_idx * batch_size
                                                end_idx = min((batch_idx + 1) * batch_size, len(values))
                                                
                                                batch_values = values[start_idx:end_idx]
                                                batch_row_indices = row_indices[start_idx:end_idx]
                                                
                                                logger.info(f"      批次 {batch_idx + 1}/{total_batches}")
                                                render_log_panel(log_panel_placeholder)
                                                
                                                extracted = ai_extract_batch(
                                                    batch_values,
                                                    f"{var_name}.{source_col}",
                                                    cache=st.session_state.ai_cache,
                                                )
                                                render_log_panel(log_panel_placeholder)
                                                
                                                for row_idx, result in zip(batch_row_indices, extracted):
                                                    ai_results[row_idx] = result
                                                
                                                if batch_idx < total_batches - 1:
                                                    time.sleep(AI_CONFIG["SLEEP_TIME"])
                                    
                                    logger.info(f"    AI提取完成，共处理 {len(ai_results)} 条数据")
                                    
                                    def apply_rules_with_ai(row):
                                        all_values = []
                                        
                                        for rule_idx, rule in enumerate(rules):
                                            cond_col = rule.get('condition_column', '')
                                            cond_op = rule.get('condition_operator', '=')
                                            cond_val = rule.get('condition_value', '')
                                            
                                            if not cond_col or cond_col not in row.index:
                                                continue
                                            
                                            if evaluate_condition(row[cond_col], cond_op, cond_val):
                                                if rule.get('extract_type') == 'AI提取':
                                                    if row.name in ai_results:
                                                        result = ai_results[row.name]
                                                        if result:
                                                            all_values.append(result)
                                                else:
                                                    extracted = extract_value(
                                                        row,
                                                        rule.get('extract_type', '直接提取'),
                                                        rule.get('extract_value_type', '从列提取'),
                                                        rule.get('extract_value', ''),
                                                        rule.get('regex_pattern', ''),
                                                        rule.get('capture_group', 1)
                                                    )
                                                    all_values.extend(extracted)
                                        
                                        if not all_values:
                                            return ''
                                        
                                        combined = separator.join(all_values)
                                        split_values = [v.strip() for v in combined.split(separator) if v.strip()]
                                        unique_sorted = sorted(set(split_values))
                                        
                                        return separator.join(unique_sorted)
                                    
                                    df[var_name] = df.apply(apply_rules_with_ai, axis=1)
                                    
                                else:
                                    if rules:
                                        logger.info(f"    使用规则提取")
                                        df[var_name] = df.apply(
                                            lambda row: process_variable_rules(row, rules, separator),
                                            axis=1
                                        )
                                        logger.info(f"    规则提取完成")
                        
                        logger.info(f"  写入Excel: {sheet_name}")
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        worksheet = writer.sheets[sheet_name]
                        
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
                        header_font = Font(bold=True)
                        header_alignment = Alignment(horizontal='center', vertical='center')
                        
                        for col_idx, col in enumerate(df.columns, 1):
                            cell = worksheet.cell(row=1, column=col_idx)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                            cell.border = thin_border
                        
                        for row_idx in range(2, len(df) + 2):
                            for col_idx in range(1, len(df.columns) + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.border = thin_border
                        
                        worksheet.freeze_panes = 'A2'
                        worksheet.auto_filter.ref = worksheet.dimensions
                        
                        logger.info(f"  工作表 {sheet_name} 格式化完成")
                
                output.seek(0)
                
                original_name = st.session_state.uploaded_file.name
                if original_name.endswith('.xlsx'):
                    new_name = original_name.replace('.xlsx', '_processed.xlsx')
                elif original_name.endswith('.xls'):
                    new_name = original_name.replace('.xls', '_processed.xlsx')
                else:
                    new_name = original_name + '_processed.xlsx'
                
                logger.info(f"文件处理完成: {new_name}")
                
                st.download_button(
                    label="⬇️ 下载处理后的文件",
                    data=output,
                    file_name=new_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                st.success("✅ 文件处理完成!")
                logger.info("=" * 80)
                logger.info("导出流程结束")
                logger.info("=" * 80)
                
            except Exception as e:
                logger.error(f"处理失败: {str(e)}", exc_info=True)
                st.error(f"❌ 处理失败: {str(e)}")
                st.exception(e)

# 页脚
st.markdown("---")
st.markdown(
    "<p style='text-align: center; color: #6b7280;'>医学编码数据预处理器 v2.1 (含AI提取) | Powered by Streamlit & DeepSeek</p>",
    unsafe_allow_html=True
)
